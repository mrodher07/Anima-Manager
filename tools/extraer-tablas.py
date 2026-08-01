#!/usr/bin/env python3
"""Extrae las tablas de reglas de la ficha Meirmeister.xlsm a JSON."""
import json, os, re, warnings
import openpyxl

warnings.filterwarnings('ignore')

SRC = os.path.join(os.path.dirname(__file__), 'ficha.xlsm')
OUT = os.environ.get('OUT_DIR', '/home/user/Anima-Manager/data/reglas')
wb = openpyxl.load_workbook(SRC, data_only=True)
# Segunda copia sin resolver: hace falta para el árbol del Ki, donde el coste vive
# dentro de la fórmula (`=IF(Q10=0,40,"-")`) y el valor resuelto es "-" en cuanto el
# personaje de la ficha ya tiene esa habilidad.
wbf = openpyxl.load_workbook(SRC, data_only=False)


def cells(sheet, ref):
    return wb[sheet][ref.replace('$', '')]


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    if isinstance(v, float) and abs(v - round(v)) < 1e-9:
        return int(round(v))
    return v


def table(sheet, ref, headers, key_idx=0, drop_headings=True):
    """Filas -> dicts. Descarta filas vacías y las cabeceras de sección '> XXX'."""
    rows, section = [], None
    for row in cells(sheet, ref):
        vals = [clean(c.value) for c in row]
        key = vals[key_idx]
        if key is None:
            continue
        if isinstance(key, str) and key.startswith('>'):
            section = key.lstrip('> ').strip()
            if drop_headings:
                continue
        item = {h: v for h, v in zip(headers, vals) if v is not None}
        if section:
            item['_seccion'] = section
        rows.append(item)
    return rows


def matrix(sheet, ref, headers):
    return table(sheet, ref, headers, drop_headings=False)


data = {}

# --- Razas -------------------------------------------------------------
RAZA_H = ['raza', 'RF', 'RE', 'RV', 'RM', 'RP', 'ajusteNivel',
          'AGI', 'CON', 'DES', 'FUE', 'INT', 'PER', 'POD', 'VOL',
          'tamano', 'regeneracion', 'cansancio', 'natura', 'descripciones']
data['razas'] = matrix('Tablas', '$J$109:$AC$129', RAZA_H)

# --- Categorías --------------------------------------------------------
CAT_H = ['categoria', 'turno', 'PV', 'costeMultiploPV', 'conocimientoMarcial',
         'limiteCombate', 'limiteMagia', 'limitePsi', 'nvPorCV',
         'bonoHA', 'bonoHP', 'bonoHE', 'bonoLlevarArmadura', 'bonoZeon',
         'costeHA', 'costeHP', 'costeHE', 'costeLlevarArmadura', 'costeKi',
         'costeAcumKi', 'costeZeon', 'costeACT', 'costeProyeccionMagica',
         'costeConvocar', 'costeControlar', 'costeAtar', 'costeDesconvocar',
         'costeCV', 'costeProyeccionPsiquica',
         'costeAtleticas', 'costeSociales', 'costePerceptivas',
         'costeIntelectuales', 'costeVigor', 'costeSubterfugio',
         'costeCreativas',
         'costePFuerza', 'costeResDolor', 'costeFrialdad', 'costeTramperia',
         'costeHerbolaria', 'costeAnimales', 'costeMedicina', 'costeTasacion',
         'costeSigilo', 'costeMemorizar', 'costeVMagica', 'costeTManos',
         'costePersuasion', 'costeOcultismo',
         'bonPFuerza', 'bonAcrobacias', 'bonSaltar', 'bonAtletismo',
         'bonTManos', 'bonEstilo', 'bonLiderazgo', 'bonResDolor',
         'bonIntimidar', 'bonFrialdad', 'bonPersuasion', 'bonDesconvocar',
         'bonControlar', 'bonAdvertir', 'bonBuscar', 'bonRastrear',
         'bonTramperia', 'bonAnimales', 'bonHerbolaria', 'bonOcultarse',
         'bonSigilo', 'bonRobo', 'bonVenenos', 'bonOcultismo', 'bonVMagica',
         'bonDisfraz', 'bonConvocar', 'bonAtar', 'bonDeteccionKi',
         'bonOcultacionKi', 'arquetipo1', 'arquetipo2', 'arquetipo']
data['categorias'] = matrix('Tablas', '$D$202:$CH$223', CAT_H)

# --- Ventajas y desventajas -------------------------------------------
VENT_H = ['nombre', 'coste', '_adq', '_pts', 'implementada', 'tipo']
ventajas = table('Tablas', '$E$256:$J$569', VENT_H)
for v in ventajas:
    v.pop('_adq', None)
    v.pop('_pts', None)
    v['esDesventaja'] = isinstance(v.get('coste'), int) and v['coste'] < 0
data['ventajas'] = ventajas

# --- Habilidades esenciales (Ventajas/Desventajas esenciales) ---------
data['habilidadesEsenciales'] = table(
    'Tablas', '$E$1249:$J$1455',
    ['nombre', 'gnosis', 'coste', '_adq', 'implementada', '_bono'])
for h in data['habilidadesEsenciales']:
    h.pop('_adq', None)
    h.pop('_bono', None)

# --- Armas y escudos ---------------------------------------------------
ARMA_H = ['arma', 'dano', 'turno', 'fueRequerida', 'fueReq2M', 'critico1',
          'critico2', 'tipoArma', 'conocida', 'entereza', 'rotura',
          'presencia', 'bonusParada', 'bonusEsquiva', 'cadencia', 'recarga',
          'alcance', 'fuerza', 'especial', '_x', 'tamano', '_atrDano']
armas = table('Tablas', '$D$638:$Y$840', ARMA_H)
for a in armas:
    a.pop('_x', None)
    # 'Atr. Daño' (col Y) no es dato del arma: en la hoja es la fórmula
    # =Bono_Fue, es decir el bono de FUE del personaje. Se descarta.
    a.pop('_atrDano', None)
data['armas'] = [a for a in armas if not str(a['arma']).startswith('Arma #')]

# --- Armaduras ---------------------------------------------------------
data['armaduras'] = table('Tablas', '$D$580:$R$632', [
    'armadura', 'requerimiento', 'penNatural', 'restMovimiento', 'entereza',
    'presencia', 'localizacion', 'clase',
    'FIL', 'CON', 'PEN', 'CAL', 'ELE', 'FRI', 'ENE'])

data['yelmos'] = table('Tablas', '$V$619:$AI$628', [
    'yelmo', 'requerimiento', 'penNatural', 'entereza', 'presencia',
    'localizacion', 'clase',
    'FIL', 'CON', 'PEN', 'CAL', 'ELE', 'FRI', 'ENE'])

# --- Artes marciales ---------------------------------------------------
data['artesMarciales'] = table('Tablas', '$D$850:$Z$939', [
    'arte', 'danoBase', 'bonoDano', 'CM', 'bonoAtaque', 'bonoEsquiva',
    'bonoParada', 'bonoTurno', 'bonoEntereza', 'bonoRotura', 'especial',
    '_o', '_p', '_q', '_r', 'longEsp', 'critico1', 'critico2',
    'bonoMaestroAt', 'bonoMaestroDef', 'danoMaximo', '_adq', 'requisitos'])
for a in data['artesMarciales']:
    for k in ('_o', '_p', '_q', '_r', '_adq'):
        a.pop(k, None)

# --- Ars Magnus --------------------------------------------------------
data['arsMagnus'] = table('Tablas', '$E$985:$J$1046', [
    'nombre', 'PD', 'CM', '_adq', 'requisitos', 'descripcion'])
for a in data['arsMagnus']:
    a.pop('_adq', None)

# --- Habilidades del Ki y del Némesis ----------------------------------
# La hoja «Ki» dibuja los dos árboles con caracteres de línea (├ └ │). El nombre de
# cada habilidad está en la columna que corresponde a su profundidad, así que la
# columna basta para saber de quién cuelga: el padre es el último nombre que hay por
# encima una columna a la izquierda.
GLIFOS = '├└│  '

# La hoja abrevia para que quepa; aquí se devuelven los nombres del manual.
NOMBRE_LARGO = {
    'Mult. de cuerpos': 'Multiplicación de cuerpos',
    'Mult. mayor': 'Multiplicación de cuerpos mayor',
    'Mult. arcana': 'Multiplicación de cuerpos arcana',
    'Mag. arcana': 'Magnitud arcana',
    'Mov. de masas': 'Movimiento de masas',
    'Armadura mayor': 'Armadura de energía mayor',
    'Arm. arcana': 'Armadura de energía arcana',
    'Inmunidad elem. FUE': 'Inmunidad elemental: Fuego',
    'Inmunidad elem. FRI': 'Inmunidad elemental: Frío',
    'Inmunidad elem. ELE': 'Inmunidad elemental: Electricidad',
}


def arbol_habilidades(hoja, filas, columnas, col_coste, dominio):
    """Recorre un árbol dibujado en columnas y devuelve nombre, requisito y coste."""
    hf, hv = wbf[hoja], wb[hoja]
    salida, ultimo_en = [], {}
    for fila in filas:
        for profundidad, col in enumerate(columnas):
            crudo = hv.cell(fila, col).value
            if not isinstance(crudo, str):
                continue
            nombre = crudo.strip(GLIFOS).strip()
            if not nombre:
                continue
            nombre = NOMBRE_LARGO.get(nombre, nombre)
            # El coste está dentro de la fórmula, no en el valor resuelto.
            formula = str(hf.cell(fila, col_coste).value or '')
            m = re.search(r',\s*(\d+)\s*,\s*"-"', formula)
            if not m:
                continue
            salida.append({
                'habilidad': nombre,
                'dominio': dominio,
                'requisito': ultimo_en.get(profundidad - 1),
                'CM': int(m.group(1)),
            })
            ultimo_en[profundidad] = nombre
            # Un nombre nuevo a esta profundidad invalida lo que colgaba más adentro.
            for mas_hondo in [d for d in ultimo_en if d > profundidad]:
                ultimo_en.pop(mas_hondo, None)
            break
    return salida


habilidades_ki = arbol_habilidades('Ki', range(10, 65), (11, 12, 13, 14), 16, 'Ki')
# El árbol del Némesis arranca en «Uso del Némesis», que no depende de nada.
habilidades_ki += arbol_habilidades('Ki', range(43, 65), (3, 4, 5), 8, 'Némesis')

# La hoja dibuja algunas ramas con `└` a la misma profundidad que sus hermanas, de
# modo que el requisito sale mal. Aquí manda el manual (Dominus Exxet, cap. 3).
REQUISITO_CORREGIDO = {
    'Multiplicación de cuerpos arcana': 'Multiplicación de cuerpos mayor',
    'Magnitud arcana': 'Magnitud',
}
# Némesis: Inhumanidad y Zen repiten el nombre de las del Ki, pero son otra cosa
# (Dominus Exxet las llama «Inhumanidad (Némesis)» y «Zen (Némesis)»).
SUFIJO_NEMESIS = {'Inhumanidad': 'Inhumanidad (Némesis)', 'Zen': 'Zen (Némesis)'}
# Raíz de cada dominio: todo lo que la hoja deja al ras cuelga de ella.
RAIZ = {'Ki': 'Uso del Ki', 'Némesis': 'Uso del Némesis'}

for h in habilidades_ki:
    if h['dominio'] == 'Némesis':
        h['habilidad'] = SUFIJO_NEMESIS.get(h['habilidad'], h['habilidad'])
        h['requisito'] = SUFIJO_NEMESIS.get(h['requisito'], h['requisito'])
    raiz = RAIZ[h['dominio']]
    if h['requisito'] is None and h['habilidad'] != raiz:
        h['requisito'] = raiz
    h['requisito'] = REQUISITO_CORREGIDO.get(h['habilidad'], h['requisito'])

# Forma de Vacío pide dos: el árbol sólo puede dibujar una.
for h in habilidades_ki:
    if h['habilidad'] == 'Forma de Vacío':
        h['requisitoExtra'] = 'Cuerpo de Vacío'
data['habilidadesKi'] = habilidades_ki

# --- Creación de Técnicas: efectos, opciones y coste -------------------
# Cada fila es una **opción** de un efecto: «Habilidad de Ataque» + «+25» cuesta 3
# puntos de Ki de la característica principal, 5 de la secundaria y 5 de CM.
data['efectosTecnica'] = table('Tablas Técnicas', '$C$10:$K$643', [
    'efecto', 'opcion', 'kiPrincipal', 'kiSecundaria', 'CM',
    'mantenimiento', 'sostenidaMenor', 'sostenidaMayor', 'nivel'])
for e in data['efectosTecnica']:
    # Un efecto tiene varias opciones; lo que identifica una fila es la pareja.
    e['referencia'] = f"{e['efecto']} {e.get('opcion', '')}".strip()
# «EFECTOS PERSONALIZADOS» son los huecos vacíos que la hoja deja para inventarse
# efectos. Quien quiera los suyos los añade desde Contenido propio.
data['efectosTecnica'] = [
    e for e in data['efectosTecnica'] if e.get('_seccion') != 'EFECTOS PERSONALIZADOS']

# Ficha de cada efecto: a qué característica va, de qué tipo y clase es, y con qué
# elementos casa. `caracteristicas` viene como «DES (AGI+2, FUE+2, POD+2, VOL+3)»:
# la primera es la principal y entre paréntesis van las alternativas con su recargo.
data['tiposEfectoTecnica'] = table('Tablas Técnicas', '$O$9:$W$98', [
    'efecto', '_ref', 'tipo', 'clase', 'caracteristicas',
    'elemento1', 'elemento2', 'elemento3', 'elementos'])
for t in data['tiposEfectoTecnica']:
    t.pop('_ref', None)
# El bloque «EFECTOS PERSONALIZADOS» son huecos vacíos de la propia hoja y, a
# continuación, la tabla de Reducción de CM, que no es un efecto. Fuera los dos: quien
# quiera inventarse efectos los añade desde Contenido propio.
data['tiposEfectoTecnica'] = [
    t for t in data['tiposEfectoTecnica']
    if t.get('_seccion') != 'EFECTOS PERSONALIZADOS']

# --- Compendio de Técnicas del Dominus Exxet --------------------------
# Las filas 655-665 son huecos para las Técnicas propias del jugador; el compendio
# publicado empieza en la 666.
data['tecnicasCompendio'] = table('Tablas Técnicas', '$C$666:$J$854', [
    'tecnica', '_arbol', 'nivel', 'CM', '_cmReducido', 'coste', 'efectos',
    'desventajas'])
for t in data['tecnicasCompendio']:
    t.pop('_arbol', None)
    t.pop('_cmReducido', None)
    t['arbol'] = t.pop('_seccion', None)

# --- Conjuros ----------------------------------------------------------
data['conjuros'] = table('Tablas Magia', '$D$6:$Z$680', [
    'conjuro', 'via', 'nivel', 'diario', 'tipo', 'accion',
    'intRBase', 'intRIntermedio', 'intRAvanzado', 'intRArcano',
    'zeonBase', 'zeonIntermedio', 'zeonAvanzado', 'zeonArcano',
    'mantBase', 'mantIntermedio', 'mantAvanzado', 'mantArcano',
    'efectoBase', 'efectoIntermedio', 'efectoAvanzado', 'efectoArcano',
    'efecto'])

# --- Poderes psíquicos y disciplinas -----------------------------------
data['poderesPsiquicos'] = table('Tablas psiquica', '$D$8:$R$145', [
    'poder', 'disciplina', 'nivel', 'mantenido', 'accion',
    'RUT', 'FAC', 'MED', 'DIF', 'MDF', 'ABS', 'CIM', 'IMP', 'INH', 'ZEN'])

data['disciplinasPsiquicas'] = table('Tablas', '$D$1190:$E$1202',
                                     ['disciplina', 'modificadores'])

# --- Poderes de criatura ----------------------------------------------
data['poderesCriatura'] = table('Tablas', '$O$1249:$R$1755',
                                ['nombre', 'gnosis', 'coste', '_adq'])
for p in data['poderesCriatura']:
    p.pop('_adq', None)

# --- Elan --------------------------------------------------------------
elan, patron = [], None
for row in cells('Tablas', '$E$1764:$L$1963'):
    v = [clean(c.value) for c in row]
    if v[0] is None:
        continue
    if v[2] == 'Elan':          # fila cabecera de bloque: nombra al patrón
        patron = v[0]
        continue
    elan.append({'patron': patron, 'nombre': v[0], 'elan': v[2],
                 'requisito': v[3], 'coste': v[4], 'descripcion': v[7]})
data['elan'] = elan

# --- Tablas numéricas base --------------------------------------------
base = {}
base['bonoCaracteristica'] = [
    {'valor': clean(r[0].value), 'bono': clean(r[1].value),
     'multiplicadorPV': clean(r[2].value)}
    for r in cells('Tablas', '$C$14:$E$33')]
# Tabla 55 del manual. Se usa tres veces con índices distintos:
#   col PV  indexada por CON -> Puntos de Vida base
#   col PV  indexada por POD -> Zeón base
#   col ACT indexada por POD -> base de Acumulación (ACT)
# La 3.ª columna NO es el Cansancio (ese sale de CON + raza).
base['valoresBase'] = [
    {'valor': clean(r[0].value), 'PV': clean(r[1].value), 'ACT': clean(r[2].value)}
    for r in cells('Tablas', '$M$37:$O$56')]
base['fuerza'] = [
    {'valor': clean(r[0].value), 'bonoTamano': clean(r[1].value),
     'pesoKg': clean(r[2].value), 'pesoMaxKg': clean(r[3].value)}
    for r in cells('Tablas', '$G$14:$J$33')]
# Tabla 53: Acumulación de Ki base. Vale para **cualquiera** de las seis
# características acumulables, no sólo para POD: 1-9 → 1, 10-12 → 2, 13-15 → 3, 16+ → 4.
base['acumulacionKi'] = [
    {'valor': clean(r[0].value), 'acumulacion': clean(r[1].value)}
    for r in cells('Tablas', '$P$14:$Q$33') if clean(r[0].value)]
# Tabla 68 del manual: Potencial Psíquico base según VOL.
base['potencialPsiquico'] = [
    {'VOL': clean(r[0].value), 'potencial': clean(r[1].value)}
    for r in cells('Tablas', '$W$1064:$X$1083') if clean(r[0].value)]
# Tabla 70: incrementar el Potencial gastando CV (CV acumulados -> bono).
base['potencialPorCV'] = [
    {'CVacumulados': clean(r[0].value), 'bono': clean(r[1].value)}
    for r in cells('Tablas', '$L$1104:$M$1113') if clean(r[0].value)]
base['gnosis'] = [
    {'gnosis': clean(r[0].value), 'PDs': clean(r[1].value), 'nivelesSobrenat': clean(r[2].value)}
    for r in cells('Tablas', '$W$27:$Y$37')]
base['limitesKi'] = [
    {'limite': clean(r[1].value), 'coste': clean(r[2].value), 'efecto': clean(r[3].value)}
    for r in cells('Tablas', '$C$1065:$M$1071') if clean(r[1].value)]
base['cordura'] = [[clean(c.value) for c in r] for r in cells('Tablas', '$AA$27:$AC$36')]
base['fama'] = [[clean(c.value) for c in r] for r in cells('Tablas', '$AE$27:$AG$34')]
base['idiomas'] = [clean(r[0].value) for r in cells('Tablas', '$AI$27:$AJ$46') if clean(r[0].value)]
base['nivelMagia'] = [[clean(c.value) for c in r] for r in cells('Tablas', '$P$1065:$Q$1084')]
base['experienciaNecesaria'] = {
    'nota': 'fila = nivel actual; columnas = ajuste de nivel 0..10',
    'filas': [[clean(c.value) for c in r] for r in cells('Tablas', '$P$69:$AA$99')]}
base['armasEnormes'] = [
    {'tamano': clean(r[0].value), 'fueMin': clean(r[1].value),
     'tamanoMin': clean(r[2].value), 'penFUE': clean(r[3].value),
     'multDano': clean(r[4].value), 'c5': clean(r[5].value), 'c6': clean(r[6].value)}
    for r in cells('Tablas', '$AE$598:$AK$600') if clean(r[0].value)]
base['tipologias'] = [[clean(c.value) for c in r] for r in cells('Tablas', '$X$603:$AA$614')]
data['tablasBase'] = base

os.makedirs(OUT, exist_ok=True)
index = {}
for name, payload in data.items():
    path = os.path.join(OUT, f'{name}.json')
    with open(path, 'w', encoding='utf-8') as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=1)
    index[name] = len(payload)
    print(f'{name:24s} {len(payload):5d} -> {path}')

with open(os.path.join(OUT, 'index.json'), 'w', encoding='utf-8') as fh:
    json.dump({'fuente': 'Meirmeister.xlsm', 'conjuntos': index}, fh,
              ensure_ascii=False, indent=1)
